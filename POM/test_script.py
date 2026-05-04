from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
from locators_POM import Locators
from login_POM import LoginPage
import os
from openpyxl import load_workbook

path = "C:/Users/mukun/Desktop/EPAM - PROGRAM/EPAM Documents/EPAM Upload/Python Program/Home_task/POM/DDT_file.xlsx"
print("Exists?", os.path.exists(path))

wb = load_workbook(path)
sheet = wb.active

driver = None
try:
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://www.saucedemo.com/")
    driver.maximize_window()

    for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
        Username = row[0]
        Password = row[1]

        login = LoginPage(driver)
        login.uname(Username)
        login.pa_word(Password)
        login.log_button()
        if login.error() > 0:
            login.close_error()
            print(f"{Username} Not able to login ISSUE")
            print(" ")
            continue
        login.verify()
        print(f"{Username} successfully logged in")
        print(" ")
        login.click_buger()
        login.lg_out()

except Exception as e:
    print(f"The Exception occured {e}")
    raise
finally:
    print("Successfully completed the Script")
    driver.quit()