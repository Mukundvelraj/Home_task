from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time



driver = None
try:
    wb = load_workbook("C:/Users/mukun/Desktop/Python_test_file/DDT_file.xlsx")
    sheet = wb.active
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://www.saucedemo.com/")
    driver.maximize_window()
    wait = WebDriverWait(driver,10)
    for rows in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only = True):
        Username = rows[0]
        Password = rows[1]
        uname = wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='user-name']")))
        uname.clear()
        uname.send_keys(Username)

        pname = wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id = 'password']")))
        pname.clear()
        pname.send_keys(Password)

        wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='login-button']"))).click()

        alert_message = driver.find_elements(By.XPATH, '//h3[@data-test="error"]')
        if len(alert_message) > 0:
            alert_close = wait.until(EC.visibility_of_element_located((By.XPATH,"//button[@class='error-button']")))
            alert_close.click()
            print(f"{Username} Not able to login")
            continue

        pro = wait.until(EC.presence_of_element_located((By.XPATH,"//span[contains(text(),'Products')]")))
        assert pro.is_displayed(),"Page is incorrect"
        print(f"{Username} successfully logged in")
        print(f"Successfully opened the product page")
        time.sleep(3)
        wait.until(EC.element_to_be_clickable((By.ID, "react-burger-menu-btn"))).click()
        time.sleep(5)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//a[@id = 'logout_sidebar_link']"))).click()
        assert "saucedemo" in driver.current_url,"Wrong page displayed"
        print(f"Successfully Completed it")
except Exception as exe:
    print(f"Error occured {exe}")
    raise
finally:
    input("Enter any button ")
    driver.quit()