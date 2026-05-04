import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time


def excel_data():
    wb = load_workbook("C:/Users/mukun/Desktop/EPAM - PROGRAM/EPAM Documents/EPAM Upload/Python Program/Home_task/Pytest/DDT_file.xlsx")
    sheet = wb.active
    test_data = []
    for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
        Username = row[0]
        Password = row[1]
        test_data.append((Username,Password))
    
    return test_data

@pytest.fixture()
def login_details(request):
    Username,Password = request.param
    return {
        "Username":Username,
        "Password":Password
    }

@pytest.mark.parametrize("login_details",excel_data(),indirect=True)
def test_browser_open(login_details):
    Username = login_details['Username']
    Password = login_details['Password']
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://www.saucedemo.com/")
    driver.maximize_window()
    wait = WebDriverWait(driver,timeout=10)
    uname = wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='user-name']")))
    uname.clear()
    uname.send_keys(Username)

    pname = wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id = 'password']")))
    pname.clear()
    pname.send_keys(Password)

    wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='login-button']"))).click()

    alert_message = driver.find_elements(By.XPATH, '//h3[@data-test="error"]')
    if len(alert_message) > 0:
        alert_close = wait.until(EC.visibility_of_element_located((By.XPATH,"//button[@class='error-button']")))
        alert_close.click()
        print(f"{Username} Not able to login")
        return

    pro = wait.until(EC.presence_of_element_located((By.XPATH,"//span[contains(text(),'Products')]")))
    assert pro.is_displayed(),"Page is incorrect"
    print(f"{Username} successfully logged in")
    print(f"Successfully opened the product page")
    time.sleep(3)
    wait.until(EC.element_to_be_clickable((By.ID, "react-burger-menu-btn"))).click()
    time.sleep(5)
    wait.until(EC.visibility_of_element_located((By.XPATH,"//a[@id = 'logout_sidebar_link']"))).click()
    assert "saucedemo" in driver.current_url,"Wrong page displayed"
    print(f"Successfully Completed it")
